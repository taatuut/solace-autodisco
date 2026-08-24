"""
End-to-end regression tests for the Solace AutoDisco pipeline.

These mirror the manual verification process used during development: they
run the actual CLI entry points (run_pipeline.py and each step script) as
subprocesses, exactly as documented in the README's Quick start section,
against a clean, isolated copy of the repo's runtime files. Nothing here
touches the real output/ directory or config.yaml.

Two things are covered:
  1. Clean-state regression of every documented command (no config.yaml /
     taxonomy_rules.yaml present — the mock demo must need zero setup).
  2. Manual taxonomy overrides: persistence across runs, merge with
     auto-derived data, and the two-tier "(Overrides)" sheets in the
     Excel report.

Run with:
    pip install -r requirements-dev.txt
    pytest
"""

import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
import yaml

REPO_ROOT = Path(__file__).resolve().parent.parent

# The subset of the repo a user actually needs to run the pipeline —
# mirrors what "git clone" would give them, minus generated/local files.
RUNTIME_FILES = [
    "src",
    "run_pipeline.py",
    "config.example.yaml",
    "output/taxonomy_rules.example.yaml",
]

BASE_SHEETS = [
    "Summary", "Message VPNs", "Queues", "Topic Catalogue",
    "Applications", "Business Objects", "Event Flows",
]
OVERRIDE_SHEETS = [
    "Applications (Overrides)", "Business Objects (Overrides)", "Event Flows (Overrides)",
]


@pytest.fixture
def project_dir(tmp_path):
    """A clean copy of the repo's runtime files, isolated from the
    developer's real output/ and config.yaml."""
    for rel in RUNTIME_FILES:
        src = REPO_ROOT / rel
        dst = tmp_path / rel
        dst.parent.mkdir(parents=True, exist_ok=True)
        if src.is_dir():
            shutil.copytree(src, dst)
        else:
            shutil.copy2(src, dst)
    return tmp_path


def run(args: list[str], cwd: Path) -> subprocess.CompletedProcess:
    """Run a documented command with the current interpreter (so it reuses
    whatever environment pytest itself is running in — no extra installs)."""
    result = subprocess.run(
        [sys.executable] + args, cwd=cwd, capture_output=True, text=True, timeout=60,
    )
    assert result.returncode == 0, (
        f"Command failed: {' '.join(args)}\n"
        f"--- stdout ---\n{result.stdout}\n--- stderr ---\n{result.stderr}"
    )
    return result


def latest(directory: Path, pattern: str) -> Path:
    matches = sorted(directory.glob(pattern), reverse=True)
    assert matches, f"No files matching {pattern!r} in {directory}"
    return matches[0]


# ---------------------------------------------------------------------------
# Clean-state regression: every command documented in the README
# ---------------------------------------------------------------------------

def test_run_pipeline_mock_needs_no_setup(project_dir):
    """README 'Demo with mock data': no config.yaml or taxonomy_rules.yaml
    should be required. Also the baseline for sheet structure with zero
    manual overrides in play."""
    assert not (project_dir / "config.yaml").exists()
    assert not (project_dir / "output" / "taxonomy_rules.yaml").exists()

    run(["run_pipeline.py", "--mock"], cwd=project_dir)

    report = latest(project_dir / "output" / "reports", "solace_autodisco_report_*.xlsx")
    wb = openpyxl.load_workbook(report)
    assert wb.sheetnames == BASE_SHEETS


def test_individual_steps(project_dir):
    """README 'Running steps individually'."""
    run(["src/semp_extractor.py", "--mock"], cwd=project_dir)
    run(["src/taxonomy_mapper.py"], cwd=project_dir)
    run(["src/report_generator.py"], cwd=project_dir)

    assert latest(project_dir / "output" / "raw", "semp_extract_*.json").exists()
    assert latest(project_dir / "output" / "reports", "mapped_data_*.json").exists()
    assert latest(project_dir / "output" / "reports", "solace_autodisco_report_*.xlsx").exists()


def test_report_generator_mock_standalone(project_dir):
    """README: report_generator.py --mock runs the full pipeline on its own.

    Regression test for a real bug found during development: this used to
    fail with `ModuleNotFoundError: No module named 'src'` when run directly
    as `python3 src/report_generator.py`, because sys.path[0] is src/ itself
    (not the repo root) — `import src.semp_extractor` couldn't resolve.
    """
    result = run(["src/report_generator.py", "--mock"], cwd=project_dir)
    assert "ModuleNotFoundError" not in result.stdout + result.stderr
    assert latest(project_dir / "output" / "reports", "solace_autodisco_report_*.xlsx").exists()


def test_live_mode_from_templates(project_dir):
    """README 'Live broker': the cp steps + --config invocation. Uses --mock
    data so this doesn't require a real broker."""
    shutil.copy2(project_dir / "config.example.yaml", project_dir / "config.yaml")
    shutil.copy2(
        project_dir / "output" / "taxonomy_rules.example.yaml",
        project_dir / "output" / "taxonomy_rules.yaml",
    )
    run(["run_pipeline.py", "--config", "config.yaml", "--mock"], cwd=project_dir)
    assert latest(project_dir / "output" / "reports", "solace_autodisco_report_*.xlsx").exists()


# ---------------------------------------------------------------------------
# Manual taxonomy overrides: persistence, merge, and the two-tier sheets
# ---------------------------------------------------------------------------

def test_manual_overrides_persist_merge_and_produce_override_sheets(project_dir):
    # First run establishes output/taxonomy_rules.yaml from auto-derived data.
    run(["run_pipeline.py", "--mock"], cwd=project_dir)
    rules_path = project_dir / "output" / "taxonomy_rules.yaml"

    with open(rules_path) as f:
        rules = yaml.safe_load(f)
    assert rules["topicPatternToBusinessObject"]["manual"] == {}
    assert rules["domainToApplicationGroup"]["manual"] == {}

    # No overrides yet -> only the 7 base sheets, no "(Overrides)" sheets.
    report = latest(project_dir / "output" / "reports", "solace_autodisco_report_*.xlsx")
    wb = openpyxl.load_workbook(report)
    assert wb.sheetnames == BASE_SHEETS

    # Add manual overrides by hand, the way a user would per the README's
    # "Manual taxonomy overrides" section.
    rules["topicPatternToBusinessObject"]["manual"] = {"acme/dev/sales/Order": "SalesOrder"}
    rules["domainToApplicationGroup"]["manual"] = {"finance": ["app_wms"]}
    with open(rules_path, "w") as f:
        yaml.dump(rules, f, default_flow_style=False, sort_keys=False)

    # Second run: manual entries must survive regeneration (persistence)
    # and be applied (merge).
    run(["run_pipeline.py", "--mock"], cwd=project_dir)

    with open(rules_path) as f:
        rules_after = yaml.safe_load(f)
    assert rules_after["topicPatternToBusinessObject"]["manual"] == {
        "acme/dev/sales/Order": "SalesOrder"
    }
    assert rules_after["domainToApplicationGroup"]["manual"] == {"finance": ["app_wms"]}
    # auto sub-keys still regenerate independently of manual.
    assert rules_after["topicPatternToBusinessObject"]["auto"]

    report = latest(project_dir / "output" / "reports", "solace_autodisco_report_*.xlsx")
    wb = openpyxl.load_workbook(report)
    assert wb.sheetnames == BASE_SHEETS + OVERRIDE_SHEETS

    # Sheet names must fit Excel's 31-character limit (openpyxl only warns,
    # Excel itself may reject or truncate — this caught a real bug where the
    # first sheet-name draft was 32-36 characters).
    for name in wb.sheetnames:
        assert len(name) <= 31, f"Sheet name too long for Excel: {name!r}"

    # Business Objects (Overrides): the overridden topic must be reclassified
    # into a new SalesOrder row, without losing the original Order row (which
    # still has other, non-overridden topics under it).
    ws = wb["Business Objects (Overrides)"]
    bos = {row[0] for row in ws.iter_rows(min_row=2, values_only=True)}
    assert "SalesOrder" in bos
    assert "Order" in bos

    # Applications (Overrides): app_wms should have gained 'finance' via the
    # domain group override, in addition to its parsed 'warehouse' domain —
    # a union, not a replacement.
    ws = wb["Applications (Overrides)"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    domain_col = header.index("Domains")
    app_col = header.index("Application ID")
    row_for_app_wms = next(
        row for row in ws.iter_rows(min_row=2, values_only=True)
        if row[app_col] == "app_wms"
    )
    domains = row_for_app_wms[domain_col].split("\n")
    assert "finance" in domains
    assert "warehouse" in domains

    # Summary sheet must report that overrides were applied.
    ws = wb["Summary"]
    summary_rows = {row[0]: row[1] for row in ws.iter_rows(values_only=True) if row[0]}
    assert summary_rows.get("Manual taxonomy overrides applied", "").startswith("Yes")


def test_old_flat_format_taxonomy_rules_file_does_not_crash(project_dir):
    """A taxonomy_rules.yaml from before the auto/manual split (flat dict,
    no 'manual' sub-key) must degrade gracefully — no manual overrides,
    not a crash."""
    run(["run_pipeline.py", "--mock"], cwd=project_dir)
    rules_path = project_dir / "output" / "taxonomy_rules.yaml"
    rules_path.write_text(
        "topicPatternToBusinessObject:\n  acme/prod/sales/Order: Order\n"
        "domainToApplicationGroup: {}\n"
    )

    run(["run_pipeline.py", "--mock"], cwd=project_dir)

    with open(rules_path) as f:
        rules_after = yaml.safe_load(f)
    assert rules_after["topicPatternToBusinessObject"]["manual"] == {}

    report = latest(project_dir / "output" / "reports", "solace_autodisco_report_*.xlsx")
    wb = openpyxl.load_workbook(report)
    assert wb.sheetnames == BASE_SHEETS
