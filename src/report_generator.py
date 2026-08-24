"""
report_generator.py — Step 3a
Generates a multi-sheet Excel report from mapped SEMP data.

Sheets:
  1. Summary          — high-level counts and metadata
  2. Message VPNs     — VPN-level configuration
  3. Queues           — all queues with stats
  4. Topic Catalogue  — unique topics parsed by taxonomy
  5. Applications     — client usernames / application identities
  6. Business Objects — BO catalogue with event types
  7. Event Flows      — matrix: which app produces/consumes which BO

Usage:
    python src/report_generator.py [--config config.yaml] [--input <mapped_data.json>] [--mock]
    python src/report_generator.py --mock   # run full pipeline (extract + map + report)
"""

import argparse
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import yaml

try:
    import openpyxl
    from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    sys.exit("[ERROR] openpyxl not installed. Run: pip install openpyxl")


# ---------------------------------------------------------------------------
# Style constants — Solace brand-adjacent palette
# ---------------------------------------------------------------------------

COLOR_HEADER_BG = "00A651"   # Solace green
COLOR_HEADER_FG = "FFFFFF"
COLOR_ALT_ROW    = "F0FAF4"
COLOR_ACCENT1    = "007B3A"   # dark green for section headers
COLOR_WARNING    = "FFF3CD"
COLOR_CRITICAL   = "F8D7DA"
COLOR_BORDER     = "CCCCCC"

FONT_HEADER = Font(name="Calibri", bold=True, color=COLOR_HEADER_FG, size=11)
FONT_TITLE  = Font(name="Calibri", bold=True, size=14, color=COLOR_ACCENT1)
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_MONO   = Font(name="Courier New", size=9)

FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER_BG)
FILL_ALT    = PatternFill("solid", fgColor=COLOR_ALT_ROW)
FILL_WARN   = PatternFill("solid", fgColor=COLOR_WARNING)
FILL_CRIT   = PatternFill("solid", fgColor=COLOR_CRITICAL)

THIN = Side(style="thin", color=COLOR_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
WRAP   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def header_row(ws, columns: list[tuple[str, int]], row: int = 1):
    """Write a styled header row. columns = [(label, width), ...]"""
    for col_idx, (label, width) in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.border = BORDER
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def data_row(ws, values: list, row: int, alt: bool = False):
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = FONT_NORMAL
        cell.border = BORDER
        cell.alignment = LEFT
        if alt:
            cell.fill = FILL_ALT


def freeze_and_filter(ws, freeze_row: int = 2):
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)
    ws.auto_filter.ref = ws.dimensions


def bool_str(val) -> str:
    if val is True or val == "true":
        return "Yes"
    if val is False or val == "false":
        return "No"
    return str(val) if val is not None else ""


def list_str(lst: list) -> str:
    return "\n".join(str(x) for x in lst) if lst else ""


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_summary(wb, mapped: dict):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Solace Event Mesh — Metadata Report"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = CENTER

    ws.row_dimensions[1].height = 30
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    meta = [
        ("Generated at", mapped.get("mappedAt", ""), "", ""),
        ("Source file", mapped.get("sourceFile", ""), "", ""),
        ("", "", "", ""),
        ("Metric", "Count", "", ""),
    ]
    row = 2
    for m in meta:
        for ci, v in enumerate(m, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = Font(name="Calibri", bold=(ci == 1), size=10)
        row += 1

    vpns = mapped.get("msgVpns", [])
    vpn_data = mapped.get("vpnData", [])
    topics = mapped.get("uniqueTopics", [])
    apps = mapped.get("applications", [])
    bos = mapped.get("businessObjectCatalogue", [])
    total_queues = sum(len(v.get("queues", [])) for v in vpn_data)
    total_clients = sum(v.get("activeConnections", 0) for v in apps)

    stats = [
        ("Message VPNs", len(vpns)),
        ("Queues (total)", total_queues),
        ("Unique topics", len(topics)),
        ("Client applications", len(apps)),
        ("Business objects", len(bos)),
        ("Active connections (runtime)", total_clients),
    ]
    header_row(ws, [("Metric", 28), ("Value", 22)], row=row)
    row += 1
    for i, (metric, val) in enumerate(stats):
        data_row(ws, [metric, val], row, alt=(i % 2 == 1))
        row += 1


def build_vpns(wb, mapped: dict):
    ws = wb.create_sheet("Message VPNs")
    cols = [
        ("VPN Name", 22), ("Enabled", 10), ("Max Spool (MB)", 16),
        ("Basic Auth", 12), ("TLS", 10),
    ]
    header_row(ws, cols)
    for i, vpn in enumerate(mapped.get("msgVpns", []), 1):
        data_row(ws, [
            vpn.get("msgVpnName"),
            bool_str(vpn.get("enabled")),
            vpn.get("maxMsgSpoolUsage"),
            bool_str(vpn.get("authenticationBasicEnabled")),
            bool_str(vpn.get("tlsEnabled")),
        ], i + 1, alt=(i % 2 == 0))
    freeze_and_filter(ws)


def build_queues(wb, mapped: dict):
    ws = wb.create_sheet("Queues")
    cols = [
        ("VPN", 16), ("Queue Name", 32), ("Access Type", 14),
        ("Owner / App", 20), ("Max Spool (MB)", 16), ("Bind Count", 12),
        ("Msg Count", 12), ("Spool Usage (B)", 16),
        ("Subscribed Topics", 60),
    ]
    header_row(ws, cols)
    row = 2
    for vpn in mapped.get("vpnData", []):
        vpn_name = vpn["msgVpnName"]
        # Build stats index
        stats_idx = {s["queueName"]: s for s in vpn.get("queueStats", [])}
        for i, q in enumerate(vpn.get("queues", [])):
            qn = q["queueName"]
            st = stats_idx.get(qn, {})
            subs = [s.get("subscriptionTopic") or s.get("topic", "")
                    for s in q.get("_subscriptions", [])]
            data_row(ws, [
                vpn_name,
                qn,
                q.get("accessType"),
                q.get("owner"),
                q.get("maxMsgSpoolUsage"),
                st.get("bindCount"),
                st.get("msgCount"),
                st.get("msgSpoolUsage"),
                list_str(subs),
            ], row, alt=(i % 2 == 0))
            ws.row_dimensions[row].height = max(15, 15 * len(subs)) if subs else 15
            row += 1
    freeze_and_filter(ws)


def build_topics(wb, mapped: dict):
    ws = wb.create_sheet("Topic Catalogue")
    cols = [
        ("Topic (raw)", 50), ("Prefix", 10), ("Environment", 14),
        ("Domain", 18), ("Business Object", 20),
        ("Event Type", 18), ("Version", 10), ("Wildcard", 10),
    ]
    header_row(ws, cols)
    for i, pt in enumerate(mapped.get("parsedTopics", []), 1):
        data_row(ws, [
            pt.get("_raw"),
            pt.get("prefix"),
            pt.get("environment"),
            pt.get("domain"),
            pt.get("businessObject"),
            pt.get("eventType"),
            pt.get("version"),
            "Yes" if pt.get("_wildcardDepth", 0) > 0 else "No",
        ], i + 1, alt=(i % 2 == 0))
    freeze_and_filter(ws)


def build_applications(wb, mapped: dict):
    ws = wb.create_sheet("Applications")
    cols = [
        ("VPN", 16), ("Application ID", 24), ("Enabled", 10),
        ("Client Profile", 20), ("ACL Profile", 20),
        ("Owned Queues", 30), ("Business Objects Consumed", 30),
        ("Domains", 22), ("Active Connections", 18),
    ]
    header_row(ws, cols)
    for i, app in enumerate(mapped.get("applications", []), 1):
        data_row(ws, [
            app.get("msgVpnName"),
            app.get("applicationId"),
            bool_str(app.get("enabled")),
            app.get("clientProfileName"),
            app.get("aclProfileName"),
            list_str(app.get("ownedQueues", [])),
            list_str(app.get("businessObjectsConsumed", [])),
            list_str(app.get("domainsInvolved", [])),
            app.get("activeConnections", 0),
        ], i + 1, alt=(i % 2 == 0))
        lines = max(len(app.get("ownedQueues", [])), 1)
        ws.row_dimensions[i + 1].height = max(15, 15 * lines)
    freeze_and_filter(ws)


def build_business_objects(wb, mapped: dict):
    ws = wb.create_sheet("Business Objects")
    cols = [
        ("Business Object", 24), ("Domain", 18),
        ("Environments", 22), ("Event Types", 30),
        ("Versions", 14), ("Topic Count", 12), ("Sample Topics", 50),
    ]
    header_row(ws, cols)
    for i, bo in enumerate(mapped.get("businessObjectCatalogue", []), 1):
        topics = bo.get("topics", [])
        data_row(ws, [
            bo.get("businessObject"),
            bo.get("domain"),
            list_str(bo.get("environments", [])),
            list_str(bo.get("eventTypes", [])),
            list_str(bo.get("versions", [])),
            len(topics),
            list_str(topics[:5]),  # show max 5 sample topics
        ], i + 1, alt=(i % 2 == 0))
        lines = max(len(bo.get("eventTypes", [])), 1)
        ws.row_dimensions[i + 1].height = max(15, 15 * lines)
    freeze_and_filter(ws)


def build_event_flows(wb, mapped: dict):
    """
    Matrix sheet: rows = applications, columns = business objects.
    Cell value: P (produces), C (consumes), P/C (both).
    Note: Without publish-side telemetry, we can only confirm consume side.
    Publisher inference requires ACL publish topics or runtime message tracking.
    """
    ws = wb.create_sheet("Event Flows")
    apps = mapped.get("applications", [])
    bos = [bo["businessObject"] for bo in mapped.get("businessObjectCatalogue", [])]

    if not bos:
        ws["A1"] = "No business objects found. Run taxonomy_mapper first."
        return

    # Header row: first col = App, then one col per BO
    ws.cell(1, 1, "Application / Business Object").font = FONT_HEADER
    ws.cell(1, 1).fill = FILL_HEADER
    ws.cell(1, 1).alignment = CENTER
    ws.column_dimensions["A"].width = 28

    for ci, bo in enumerate(bos, 2):
        c = ws.cell(1, ci, bo)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = CENTER
        ws.column_dimensions[get_column_letter(ci)].width = 18

    # Data rows
    for ri, app in enumerate(apps, 2):
        app_id = app.get("applicationId", "")
        consumed = set(app.get("businessObjectsConsumed", []))
        # Publisher detection: currently unknown without ACL topic introspection
        # Placeholder — in Phase 2 this would use ACL publish exceptions or EP data
        produced = set()  # TODO: infer from ACL publish exceptions

        ws.cell(ri, 1, app_id).font = FONT_NORMAL

        for ci, bo in enumerate(bos, 2):
            p = bo in produced
            c = bo in consumed
            val = ""
            fill = None
            if p and c:
                val = "P / C"
                fill = PatternFill("solid", fgColor="C8E6C9")
            elif p:
                val = "P"
                fill = PatternFill("solid", fgColor="BBDEFB")
            elif c:
                val = "C"
                fill = PatternFill("solid", fgColor="FFF9C4")

            cell = ws.cell(ri, ci, val)
            cell.font = FONT_NORMAL
            cell.alignment = CENTER
            if fill:
                cell.fill = fill
            cell.border = BORDER

    # Legend
    legend_row = len(apps) + 4
    ws.cell(legend_row, 1, "Legend:").font = Font(bold=True)
    ws.cell(legend_row + 1, 1, "P = Produces (publishes to topic)")
    ws.cell(legend_row + 2, 1, "C = Consumes (subscribes via queue)")
    ws.cell(legend_row + 3, 1, "P/C = Both producer and consumer")
    ws.cell(legend_row + 4, 1,
            "Note: Producer detection requires ACL publish exception data or Event Portal. "
            "Currently only consumer side (queue subscriptions) is shown.")

    ws.freeze_panes = ws.cell(2, 2)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def generate_report(mapped: dict, out_path: Path):
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_summary(wb, mapped)
    build_vpns(wb, mapped)
    build_queues(wb, mapped)
    build_topics(wb, mapped)
    build_applications(wb, mapped)
    build_business_objects(wb, mapped)
    build_event_flows(wb, mapped)

    wb.save(out_path)
    print(f"Report saved: {out_path}")


def main():
    arg_parser = argparse.ArgumentParser(
        description="Generate Excel report from mapped Solace SEMP data"
    )
    arg_parser.add_argument("--config", default="config.yaml")
    arg_parser.add_argument("--input", default=None,
                            help="Path to mapped_data JSON. "
                                 "If omitted, uses most recent file in output/reports/")
    arg_parser.add_argument("--mock", action="store_true",
                            help="Run full pipeline (extract --mock + map + report)")
    arg_parser.add_argument("--output", default=None,
                            help="Output .xlsx file path")
    args = arg_parser.parse_args()

    if args.mock:
        # Run extract then map in-process
        print("Running SEMP extractor (mock)...")
        import src.semp_extractor as ex
        sys.argv = ["semp_extractor", "--mock"]
        extract_file = ex.main()

        print("Running taxonomy mapper...")
        import importlib, types
        import src.taxonomy_mapper as tm
        sys.argv = ["taxonomy_mapper", "--input", extract_file]
        mapped_file = tm.main()
    elif args.input:
        mapped_file = args.input
    else:
        report_dir = Path("output/reports")
        candidates = sorted(report_dir.glob("mapped_data_*.json"), reverse=True)
        if not candidates:
            print("[ERROR] No mapped_data_*.json found. Run taxonomy_mapper first "
                  "or use --mock.")
            raise SystemExit(1)
        mapped_file = str(candidates[0])

    print(f"Loading: {mapped_file}")
    with open(mapped_file) as f:
        mapped = json.load(f)

    # Output path
    if args.output:
        out_path = Path(args.output)
    else:
        cfg_path = Path(args.config)
        if cfg_path.exists():
            with open(cfg_path) as f:
                config = yaml.safe_load(f)
            out_dir = Path(config.get("output", {}).get("report_dir", "output/reports"))
        else:
            out_dir = Path("output/reports")
        out_dir.mkdir(parents=True, exist_ok=True)
        ts_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = out_dir / f"solace_autodisco_report_{ts_str}.xlsx"

    generate_report(mapped, out_path)
    return str(out_path)


if __name__ == "__main__":
    main()
